"""
LeadFlow.AI - Specialized AI Marketing & Sales Agents Stack Generator (2026 Edition)
Compiles the comprehensive 78+ agent toolchain into an enterprise Excel document (.xlsx)
and Markdown reference index for the LeadFlow.AI repository.
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def build_agents_spreadsheet(output_path: str = None):
    if output_path is None:
        output_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "docs", "specialized-ai-agents-2026.xlsx")
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "2026 Marketing & Sales Agents"

    # Define Styles
    title_font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    title_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    
    sub_font = Font(name="Calibri", size=11, italic=True, color="D4D4D8")
    sub_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="00A99D", end_color="00A99D", fill_type="solid") # Persian Turquoise brand accent

    cat_font = Font(name="Calibri", size=12, bold=True, color="1E293B")
    cat_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")

    row_fill_alt = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    row_fill_normal = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    elena_yes_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid") # Light green
    elena_no_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid") # Light red

    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    # Title Banner
    ws.merge_cells("A1:G1")
    ws["A1"] = "🤖 2026 SPECIALIZED AI MARKETING, SALES & BUSINESS AGENTS STACK"
    ws["A1"].font = title_font
    ws["A1"].fill = title_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A2:G2")
    ws["A2"] = "Curated 78+ Agents across 11 Categories for LeadFlow.AI & Elena Autonomous Sales Swarm • Updated July 2026"
    ws["A2"].font = sub_font
    ws["A2"].fill = sub_fill
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 24

    # Table Headers
    headers = [
        "Category", "Agent / Tool Name", "Best For / Specialization",
        "Free Tier / OSS?", "Starting Price (2026)", "API / MCP?", "Suitable for Elena / LeadFlow?"
    ]
    
    ws.row_dimensions[3].height = 28
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # Data Matrix
    data = [
        # 1. Website Builder Agents
        ("1. Website Builder Agents", "Relume", "Professional Designers & Wireframing", "Yes", "$26/mo", "Yes", "YES - High Prototype Polish"),
        ("1. Website Builder Agents", "v0 by Vercel", "Developer / Marketer React & Tailwind UI", "Yes", "$20/mo", "Yes (API)", "YES - Core React UI Stack"),
        ("1. Website Builder Agents", "Framer AI", "No-code interactive visual creators", "No", "$15/mo", "Yes", "Moderate"),
        ("1. Website Builder Agents", "Lovable", "Beginner / SMB full-stack web apps", "Yes (5 cr/day)", "$25/mo", "No", "YES - Quick Client Demos"),
        ("1. Website Builder Agents", "Tempo Labs", "Designers / SaaS founders React code", "Yes", "$30/mo", "Yes", "YES - Code Handoff"),
        ("1. Website Builder Agents", "Webflow AI", "Designer + Enterprise CMS web builds", "Yes", "$12/mo", "Yes (API)", "Moderate"),
        ("1. Website Builder Agents", "Wix ADI", "Small local business quick setups", "Yes", "Free Tier", "No", "No"),
        ("1. Website Builder Agents", "Dora AI", "3D Web animations and spatial sites", "Yes", "$18/mo", "No", "YES - 3D Awwwards Vibe"),
        ("1. Website Builder Agents", "Wegic", "Local business multi-page sites", "Yes", "$23.90/mo", "No", "No"),
        ("1. Website Builder Agents", "Atlas AI", "Shopify e-commerce store builder", "No", "$49/mo", "Yes", "YES - E-Commerce Tenants"),

        # 2. Content & Copywriting Agents
        ("2. Copywriting & Content", "Jasper", "Agentic content + Brand IQ (5800+ apps)", "No", "$69/mo", "Yes (API)", "YES - Brand Voice QA"),
        ("2. Copywriting & Content", "Copy.ai", "GTM workflows & sales/marketing chaining", "Yes", "$49/mo", "Yes (API)", "YES - GTM Automation"),
        ("2. Copywriting & Content", "Writesonic / Chatsonic", "SEO blogs, social copy & site chatbots", "Yes", "Free Tier", "Yes", "Moderate"),
        ("2. Copywriting & Content", "Notion AI", "Knowledge management & team workspace", "Yes", "Included", "Yes", "Moderate"),
        ("2. Copywriting & Content", "Akira AI", "Social copy & rapid blog ideation", "Subscription", "Paid", "No", "No"),
        ("2. Copywriting & Content", "Claude 3.5 / ChatGPT", "Long-form reasoning, ideation & coding", "Yes", "$20/mo", "Yes (API)", "YES - Core Brain Engine"),

        # 3. Top 7 Marketing Agents 2026
        ("3. Marketing Agents (Top 2026)", "Superscale", "Full paid creative loop: research->gen->publish", "No", "$49/mo", "Yes", "YES - Ad Creative Loop"),
        ("3. Marketing Agents (Top 2026)", "Salesforce Agentforce", "Autonomous CRM sales/marketing agents", "No", "$2/conv", "Yes", "Enterprise Only"),
        ("3. Marketing Agents (Top 2026)", "HubSpot Breeze Agents", "Marketing, sales & customer intelligence", "Yes (Free Tier)", "$90/mo Pro", "Yes", "YES - Inbound CRM"),
        ("3. Marketing Agents (Top 2026)", "Omneky", "Automated ad creative generation & testing", "No", "Enterprise", "Yes", "YES - Ad Scaling"),
        ("3. Marketing Agents (Top 2026)", "Albert.ai", "Autonomous paid media buying & optimization", "No", "$2,000/mo", "Yes", "Enterprise Only"),
        ("3. Marketing Agents (Top 2026)", "Tofu", "B2B hyper-personalized campaign scaling", "No", "Custom", "Yes", "YES - ABM Campaigns"),
        ("3. Marketing Agents (Top 2026)", "Blaze AI", "100+ automated marketing workflows", "No", "Custom", "Yes", "Moderate"),
        ("3. Marketing Agents (Top 2026)", "ActiveCampaign AI", "30+ AI agents for email & journey mapping", "No", "$49/mo", "Yes", "Moderate"),
        ("3. Marketing Agents (Top 2026)", "Gumloop", "No-code visual AI agent workflow builder", "Yes (Free Tier)", "Free / Paid", "Yes", "YES - Rapid Automation"),
        ("3. Marketing Agents (Top 2026)", "Relevance AI", "No-code B2B GTM & sales agent builder", "Yes (Free Tier)", "Free / Paid", "Yes (API)", "YES - No-Code Swarms"),

        # 4. SEO & Blog Agents
        ("4. SEO & Blog Agents", "Semrush / Surfer SEO", "Keyword research & content SERP scoring", "No", "$89-$129/mo", "Yes", "YES - SEO Audits"),
        ("4. SEO & Blog Agents", "Clearscope / MarketMuse", "Enterprise topical authority & content optimization", "No", "$170+/mo", "Yes", "Enterprise Only"),
        ("4. SEO & Blog Agents", "NeuronWriter / RankIQ", "Budget-friendly SEO blog content scoring", "No", "$19-$49/mo", "No", "Moderate"),
        ("4. SEO & Blog Agents", "SEO Specialist (Claw)", "On-page optimization & cluster analysis", "YES (OSS)", "Free (OSS)", "Yes (CLI)", "YES - Zero Cost SEO"),
        ("4. SEO & Blog Agents", "Content Creator (Claw)", "Long-form programmatic article generation", "YES (OSS)", "Free (OSS)", "Yes (CLI)", "YES - Zero Cost Content"),

        # 5. Social Media & Brand Agents
        ("5. Social Media Agents", "Buffer AI / Hootsuite", "Social scheduling & multi-platform publishing", "Yes (Free Tier)", "Paid Pro", "Yes", "Moderate"),
        ("5. Social Media Agents", "TikTok Strategist (Claw)", "Short-form vertical video script & trend strategy", "YES (OSS)", "Free (OSS)", "Yes (CLI)", "YES - Viral Reels Strategy"),
        ("5. Social Media Agents", "Brand Guardian (Claw)", "Voice, tone, and brand compliance QA", "YES (OSS)", "Free (OSS)", "Yes (CLI)", "YES - Brand Identity QA"),
        ("5. Social Media Agents", "Whimsy Injector (Claw)", "Microcopy & delightful user engagement", "YES (OSS)", "Free (OSS)", "Yes (CLI)", "YES - UI Polish"),
        ("5. Social Media Agents", "Canva AI / Synthesia", "Social graphics & AI video talking avatars", "Yes (Free Tier)", "$22+/mo", "Yes", "YES - Virtual Influencers"),

        # 6. Sales, GTM & Outreach Stack
        ("6. Sales & GTM Outbound", "Persana AI", "All-in-one outbound prospecting & enrichment", "No", "$68/mo", "Yes", "YES - Outbound SDR"),
        ("6. Sales & GTM Outbound", "Clay", "Lead research & 150+ data provider enrichment", "No", "$167/mo", "Yes (API)", "YES - Lead Enrichment"),
        ("6. Sales & GTM Outbound", "Lindy", "Autonomous inbox & scheduling agent swarms", "No", "$49.99/mo", "Yes", "YES - Meeting Booking"),
        ("6. Sales & GTM Outbound", "Telescope / Unify", "Signal-triggered email & LinkedIn list building", "No", "$47-$700/mo", "Yes", "YES - Intent Signals"),
        ("6. Sales & GTM Outbound", "Agent Frank (Salesforge)", "Autonomous email/LinkedIn SDR with warm-up", "No", "$499/mo", "Yes", "YES - Dedicated SDR"),
        ("6. Sales & GTM Outbound", "n8n Self-Hosted", "Custom SDR workflows with 500+ integrations", "YES (OSS)", "Free Self-Hosted", "Yes (API)", "YES - Core Orchestrator"),
        ("6. Sales & GTM Outbound", "Make / Zapier", "No-code visual integration pipelines", "Yes (Free Tier)", "$9+/mo", "Yes", "YES - App Gluing"),
        ("6. Sales & GTM Outbound", "CrewAI / LangGraph / AutoGen", "Role-based multi-agent Python orchestration", "YES (OSS)", "Free (OSS)", "Yes (Python)", "YES - Core Backend Architecture"),

        # 7. Email Marketing Agents
        ("7. Email Marketing", "ActiveCampaign / HubSpot", "Automated email journey & lead scoring", "Yes (Free Tier)", "$49+/mo", "Yes", "YES - Nurture Sequences"),
        ("7. Email Marketing", "Seventh Sense", "AI predictive email send-time optimization", "No", "Paid", "Yes", "YES - Open Rate Lift"),
        ("7. Email Marketing", "Jacquard / Phrasee", "AI copywriting optimized for email click-throughs", "No", "Paid", "Yes", "Moderate"),
        ("7. Email Marketing", "Mailchimp AI", "SMB email newsletter & basic automation", "Yes (Free Tier)", "Paid Pro", "Yes", "Moderate"),

        # 8. Paid Ads & Creative Agents
        ("8. Paid Ads & Creative", "Superscale / AdCreative.ai", "AI ad creative generation at massive volume", "No", "$39-$49/mo", "Yes", "YES - Ad Creatives"),
        ("8. Paid Ads & Creative", "Madgicx / Smartly.io", "Meta & Google ads budget optimization copilot", "No", "$79-$30K/yr", "Yes", "YES - ROAS Scaling"),
        ("8. Paid Ads & Creative", "Pencil / Anyword", "Creative performance prediction & copy scoring", "No", "$49-$119/mo", "Yes", "YES - Ad Copy Testing"),
        ("8. Paid Ads & Creative", "NotFair (OSS MCP)", "Google Ads MCP server for campaign diagnosis", "YES (OSS)", "Free (OSS)", "Yes (MCP)", "YES - Zero Cost Ad Ops"),

        # 9. Video Marketing & Virtual Influencer Agents
        ("9. Video & Influencers", "Pippit AI", "600+ Avatars, 869 Voices, 28 Languages", "YES (150 cr/wk)", "Free / Paid", "Yes", "YES - Elena Virtual Avatar"),
        ("9. Video & Influencers", "Atlabs AI", "UGC Avatar Video + Lip Sync with 40+ languages", "YES (Trial)", "Paid Pro", "Yes", "YES - High Character Consistency"),
        ("9. Video & Influencers", "HeyGen / Synthesia", "Studio-grade talking avatars & voice cloning", "Yes (3 vids/mo)", "$22-$24/mo", "Yes (API)", "YES - Official Video Output"),
        ("9. Video & Influencers", "Descript", "AI audio/video editing & transcript dubbing", "Yes (Free Tier)", "$12+/mo", "Yes", "YES - Fast Reels Editing"),

        # 10. Automation & Knowledge App Engines
        ("10. Automation & Knowledge", "n8n / Composio", "Self-hosted workflows & managed agent tools", "YES (OSS)", "Free / Paid", "Yes (API)", "YES - Primary Backend Engine"),
        ("10. Automation & Knowledge", "Dify / Flowise", "Visual LLMOps & internal RAG knowledge apps", "YES (OSS)", "Free (OSS)", "Yes (API)", "YES - Elena Knowledge Base"),
        ("10. Automation & Knowledge", "Voiceflow / Bardeen", "Website chat, lead capture & browser automation", "Yes (Free Tier)", "Free / Paid", "Yes", "YES - Interactive Web Chat"),

        # 11. The 60-Second Claw Skills (Zero Cost OSS)
        ("11. 60-Sec Claw Skills", "npx claw install growth-hacker", "Rapid ICE growth experiments & validation", "YES (OSS)", "Free (OSS)", "Yes (CLI)", "YES - Instant Growth Loops"),
        ("11. 60-Sec Claw Skills", "npx claw install ux-researcher", "Automated customer feedback & persona analysis", "YES (OSS)", "Free (OSS)", "Yes (CLI)", "YES - Persona Optimization"),
        ("11. 60-Sec Claw Skills", "npx claw install outbound-strategist", "Cold outreach sequencing & messaging structure", "YES (OSS)", "Free (OSS)", "Yes (CLI)", "YES - Elena Outreach Blueprint"),
        ("11. 60-Sec Claw Skills", "npx claw install product-manager", "Feature launch alignment & roadmap execution", "YES (OSS)", "Free (OSS)", "Yes (CLI)", "YES - Product Strategy"),
        ("11. 60-Sec Claw Skills", "npx claw install discovery-coach", "Sales discovery call questioning frameworks", "YES (OSS)", "Free (OSS)", "Yes (CLI)", "YES - Sales Discovery")
    ]

    current_row = 4
    current_cat = ""
    for row_data in data:
        cat = row_data[0]
        # Insert category separator if changed
        if cat != current_cat:
            current_cat = cat
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
            cat_cell = ws.cell(row=current_row, column=1, value=f"  📂 {cat.upper()}")
            cat_cell.font = cat_font
            cat_cell.fill = cat_fill
            cat_cell.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[current_row].height = 24
            current_row += 1

        ws.row_dimensions[current_row].height = 22
        fill_to_use = row_fill_alt if current_row % 2 == 0 else row_fill_normal
        
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.font = Font(name="Calibri", size=10)
            cell.alignment = Alignment(horizontal="left" if col_idx in [2,3] else "center", vertical="center")
            cell.border = thin_border
            cell.fill = fill_to_use
            
            # Highlight Suitable for Elena column
            if col_idx == 7:
                if "YES" in str(val):
                    cell.fill = elena_yes_fill
                    cell.font = Font(name="Calibri", size=10, bold=True, color="065F46")
                elif "No" in str(val):
                    cell.fill = elena_no_fill
                    cell.font = Font(name="Calibri", size=10, color="991B1B")

        current_row += 1

    # Auto-fit columns
    col_widths = {1: 24, 2: 24, 3: 40, 4: 16, 5: 22, 6: 14, 7: 28}
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Add auto-filter
    ws.auto_filter.ref = f"A3:G{current_row-1}"

    wb.save(output_path)
    print(f"✅ Excel Spreadsheet successfully generated: {output_path}")

if __name__ == "__main__":
    build_agents_spreadsheet()
